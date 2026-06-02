from typing import Any

from django.contrib import admin
from django.db.models.query import QuerySet
from django.http import HttpRequest, HttpResponse
from django.urls import reverse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _

from .models import User, Brand, Model, Car, CarPhoto, Favorite, ForumPost
from .resources import CarResource
from import_export.admin import ImportExportModelAdmin
from import_export.formats import base_formats
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class CarPhotoInline(admin.TabularInline):
    model = CarPhoto
    extra = 1
    readonly_fields = ('created_at',)
    fields = ('image_url', 'is_main', 'created_at')


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ('username', 'email', 'phone', 'role', 'last_login')
    list_filter = ('role', 'is_staff', 'is_active')
    search_fields = ('username', 'email', 'phone')
    readonly_fields = ('last_login',)
    date_hierarchy = 'last_login'
    raw_id_fields = ('groups',)


@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
    list_display = ('name', 'created_at')
    search_fields = ('name',)
    date_hierarchy = 'created_at'


@admin.register(Model)
class ModelAdmin(admin.ModelAdmin):
    list_display = ('name', 'brand', 'created_at')
    list_filter = ('brand',)
    search_fields = ('name', 'brand__name')
    date_hierarchy = 'created_at'


@admin.register(Car)
class CarAdmin(ImportExportModelAdmin):
    resource_class = CarResource

    list_display = ('id', 'full_name', 'year', 'price_formatted', 'status', 'views', 'created_at')
    list_filter = ('status', 'brand', 'year', 'created_at')
    search_fields = ('description', 'brand__name', 'model__name')
    readonly_fields = ('views', 'created_at', 'updated_at')
    date_hierarchy = 'created_at'
    inlines = [CarPhotoInline]
    raw_id_fields = ('user', 'created_by')
    list_display_links = ('id', 'full_name')
    actions = ['export_admin_action']

    def get_export_formats(self) -> list[Any]:
        """Возвращает доступные форматы экспорта для действия в админке."""
        return [
            base_formats.XLSX,
            base_formats.CSV,
            base_formats.JSON,
        ]

    def export_admin_action(self, request: HttpRequest, queryset: QuerySet[Car]) -> HttpResponse:
        """Кастомное действие экспорта в XLSX с форматированием (в админке)."""
        resource = self.resource_class()
        dataset = resource.export(queryset)

        wb = Workbook()
        ws = wb.active
        ws.title = "Объявления"

        # Жирные заголовки
        bold_font = Font(bold=True, size=12)
        for col_num, column_title in enumerate(dataset.headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center')

        # Данные
        for row_num, row in enumerate(dataset.dict, 2):
            for col_num, value in enumerate(row.values(), 1):
                ws.cell(row=row_num, column=col_num).value = value

        # Автоширина столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Заморозка первой строки
        ws.freeze_panes = 'A2'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Объявления_CarHub.xlsx"'
        wb.save(response)
        return response

    @admin.display(description=_('Полное название'))
    def full_name(self, obj: Car) -> str:
        """Отображает полное название авто в списке админки."""
        return f'{obj.brand} {obj.model} ({obj.year})'

    @admin.display(description=_('Цена'))
    def price_formatted(self, obj: Car) -> str:
        """Форматирует цену с валютой в админке."""
        return f'{obj.price} ₽'
    price_formatted.short_description = _('Цена')


@admin.register(CarPhoto)
class CarPhotoAdmin(admin.ModelAdmin):
    list_display = ('car', 'is_main', 'created_at')
    list_filter = ('is_main', 'created_at')
    search_fields = ('car__brand__name', 'car__model__name')
    readonly_fields = ('created_at',)
    date_hierarchy = 'created_at'


class FavoriteInline(admin.TabularInline):
    model = Favorite
    extra = 0
    raw_id_fields = ('car',)


@admin.register(Favorite)
class FavoriteAdmin(admin.ModelAdmin):
    list_display = ('user', 'car_link', 'created_at')
    list_filter = ('created_at',)
    search_fields = ('user__username', 'car__brand__name')
    date_hierarchy = 'created_at'
    raw_id_fields = ('user', 'car')

    @admin.display(description=_('Объявление'))
    def car_link(self, obj: Favorite) -> str:
        """Link to car in admin for Favorite."""
        if obj.car_id:
            try:
                url = reverse('admin:core_car_change', args=[obj.car_id])
                return format_html('<a href="{}">{}</a>', url, str(obj.car))
            except Exception:
                pass
        return str(obj.car) or '—'
    car_link.short_description = _('Автомобиль')
